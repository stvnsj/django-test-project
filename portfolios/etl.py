
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path


from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from portfolios.models import Asset, AssetPrice, Portfolio, PortfolioAsset

INITIAL_PORTFOLIO_VALUE = Decimal("1000000000")
WEIGHTS_SHEET_NAME = "weights"
PRICES_SHEET_NAME = "Precios"

def to_decimal(value):
    if value is None:
        raise ValueError("Expected a numeric value, got None")

    return Decimal(str(value))

def to_date(value):

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        return from_excel(value).date()

    raise ValueError(f"Could not convert value to date: {value!r}")


def normalize_asset_name(value):
    if value is None:
        raise ValueError("Asset name cannot be empty")

    return str(value).strip()

def create_portfolios(start_date):
    
    portfolio_1, _ = Portfolio.objects.update_or_create(
        name="Portfolio 1",
        defaults={
            "initial_value": INITIAL_PORTFOLIO_VALUE,
            "start_date": start_date
        }
    )

    portfolio_2, _ = Portfolio.objects.update_or_create(
        name = "Portfolio 2",
        defaults={
            "initial_value": INITIAL_PORTFOLIO_VALUE,
            "start_date": start_date
        }
    )

    return {
        "portfolio_1": portfolio_1,
        "portfolio_2": portfolio_2
    }


def load_assets_from_prices_sheet(prices_sheet):
    """
    
    
    Create assets from the header row of the prices sheet.

    """
    assets = {}

    for column in range(2, prices_sheet.max_column + 1):
        asset_name = normalize_asset_name(
            prices_sheet.cell(row=1, column=column).value
        )

        asset, _ = Asset.objects.get_or_create(name=asset_name)
        assets[asset_name] = asset

    return assets

def load_prices(prices_sheet, assets):
    """
    Create or update daily prices for all assets.
    """
    for row in range(2, prices_sheet.max_row + 1):
        current_date = to_date(prices_sheet.cell(row=row, column=1).value)
        for column in range(2, prices_sheet.max_column + 1):
            asset_name = normalize_asset_name(
                prices_sheet.cell(row=1, column=column).value
            )
            raw_price = prices_sheet.cell(row=row, column=column).value
            if raw_price is None:
                continue
            
            AssetPrice.objects.update_or_create(
                asset=assets[asset_name],
                date=current_date,
                defaults={
                    "price": to_decimal(raw_price)
                }
            )


            
def load_initial_weights(weights_sheet, assets, portfolios):
    for row in range(2, weights_sheet.max_row + 1):
        raw_asset_name = weights_sheet.cell(row=row, column=2).value

        if raw_asset_name is None:
            continue

        asset_name = normalize_asset_name(raw_asset_name)

        if asset_name not in assets:
            raise ValueError(
                f"Asset {asset_name!r} appears in weights sheet "
                "but not in prices sheet"
            )

        asset = assets[asset_name]

        portfolio_1_weight = to_decimal(weights_sheet.cell(row=row, column=3).value)
        portfolio_2_weight = to_decimal(weights_sheet.cell(row=row, column=4).value)

        PortfolioAsset.objects.update_or_create(
            portfolio=portfolios["portfolio_1"],
            asset=asset,
            defaults={
                "initial_weight": portfolio_1_weight,
            },
        )

        PortfolioAsset.objects.update_or_create(
            portfolio=portfolios["portfolio_2"],
            asset=asset,
            defaults={
                "initial_weight": portfolio_2_weight,
            },
        )


def calculate_initial_quantities():

    portfolio_assets = PortfolioAsset.objects.select_related(
        "portfolio",
        "asset",
    )

    for portfolio_asset in portfolio_assets:
        initial_price = AssetPrice.objects.get(
            asset=portfolio_asset.asset,
            date=portfolio_asset.portfolio.start_date,
        )

        if initial_price.price == 0:
            raise ValueError(
                f"Initial price cannot be zero for asset {portfolio_asset.asset}"
            )

        initial_quantity = (
            portfolio_asset.initial_weight
            * portfolio_asset.portfolio.initial_value
            / initial_price.price
        )

        portfolio_asset.initial_quantity = initial_quantity
        portfolio_asset.save(update_fields=["initial_quantity"])



    

@transaction.atomic
def load_portfolio_data(file_path):
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    workbook = load_workbook(path, data_only=True)

    weights_sheet = workbook[WEIGHTS_SHEET_NAME]
    prices_sheet = workbook[PRICES_SHEET_NAME]

    start_date = to_date(weights_sheet.cell(row=2, column=1).value)

    portfolios = create_portfolios(start_date)
    assets = load_assets_from_prices_sheet(prices_sheet)

    load_prices(prices_sheet, assets)
    load_initial_weights(weights_sheet, assets, portfolios)
    calculate_initial_quantities()
